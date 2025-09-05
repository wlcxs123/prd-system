# -*- coding: utf-8 -*-
"""
数据导出工具模块
支持 CSV、Excel、PDF 格式的问卷数据导出
"""

import json
import csv
from io import StringIO, BytesIO
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os

class QuestionnaireExporter:
    """问卷数据导出器"""
    
    def __init__(self):
        self.setup_pdf_fonts()
    
    def setup_pdf_fonts(self):
        """设置PDF中文字体支持"""
        try:
            # 尝试注册中文字体
            font_path = os.path.join(os.path.dirname(__file__), 'fonts', 'SimHei.ttf')
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont('SimHei', font_path))
                self.pdf_font = 'SimHei'
            else:
                # 如果没有中文字体文件，使用默认字体
                self.pdf_font = 'Helvetica'
        except Exception as e:
            print(f"字体设置失败，使用默认字体: {e}")
            self.pdf_font = 'Helvetica'
    
    def export_to_csv(self, questionnaires, include_details=True):
        """导出为CSV格式"""
        output = StringIO()
        writer = csv.writer(output)
        
        if not questionnaires:
            writer.writerow(['没有数据'])
            output.seek(0)
            return output.getvalue()
        
        # 写入表头
        headers = ['ID', '问卷类型', '姓名', '年级', '提交日期', '创建时间']
        if include_details:
            headers.extend(['问题总数', '完成率', '总分'])
        
        writer.writerow(headers)
        
        # 写入数据
        for q in questionnaires:
            try:
                data = json.loads(q['data']) if isinstance(q['data'], str) else q['data']
                
                row = [
                    q['id'],
                    q['type'],
                    q['name'] or '',
                    q['grade'] or '',
                    q['submission_date'] or '',
                    q['created_at'] or ''
                ]
                
                if include_details:
                    # 计算统计信息
                    questions = data.get('questions', [])
                    stats = data.get('statistics', {})
                    
                    question_count = len(questions)
                    completion_rate = stats.get('completion_rate', 0)
                    total_score = stats.get('total_score', 0)
                    
                    row.extend([question_count, f"{completion_rate}%", total_score])
                
                writer.writerow(row)
                
            except Exception as e:
                print(f"处理问卷 {q['id']} 时出错: {e}")
                # 写入基本信息，即使详细信息处理失败
                row = [q['id'], q['type'], q['name'] or '', q['grade'] or '', 
                       q['submission_date'] or '', q['created_at'] or '']
                if include_details:
                    row.extend(['错误', '错误', '错误'])
                writer.writerow(row)
        
        output.seek(0)
        return output.getvalue()
    
    def export_to_excel(self, questionnaires, include_details=True):
        """导出为Excel格式"""
        # 创建工作簿
        wb = Workbook()
        
        # 删除默认工作表
        wb.remove(wb.active)
        
        # 创建概览工作表
        ws_summary = wb.create_sheet("问卷概览")
        self._create_summary_sheet(ws_summary, questionnaires, include_details)
        
        # 如果包含详细信息，为每个问卷类型创建单独的工作表
        if include_details and questionnaires:
            questionnaire_types = {}
            
            # 按类型分组
            for q in questionnaires:
                q_type = q['type']
                if q_type not in questionnaire_types:
                    questionnaire_types[q_type] = []
                questionnaire_types[q_type].append(q)
            
            # 为每种类型创建工作表
            for q_type, type_questionnaires in questionnaire_types.items():
                if len(type_questionnaires) > 0:
                    # 限制工作表名称长度
                    sheet_name = q_type[:30] if len(q_type) > 30 else q_type
                    ws_detail = wb.create_sheet(f"详情-{sheet_name}")
                    self._create_detail_sheet(ws_detail, type_questionnaires)
        
        # 保存到BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def _create_summary_sheet(self, worksheet, questionnaires, include_details):
        """创建概览工作表"""
        # 设置标题
        worksheet.title = "问卷概览"
        
        # 写入表头
        headers = ['ID', '问卷类型', '姓名', '年级', '提交日期', '创建时间']
        if include_details:
            headers.extend(['问题总数', '完成率', '总分'])
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入数据
        for row_idx, q in enumerate(questionnaires, 2):
            try:
                data = json.loads(q['data']) if isinstance(q['data'], str) else q['data']
                
                values = [
                    q['id'],
                    q['type'],
                    q['name'] or '',
                    q['grade'] or '',
                    q['submission_date'] or '',
                    q['created_at'] or ''
                ]
                
                if include_details:
                    questions = data.get('questions', [])
                    stats = data.get('statistics', {})
                    
                    question_count = len(questions)
                    completion_rate = stats.get('completion_rate', 0)
                    total_score = stats.get('total_score', 0)
                    
                    values.extend([question_count, f"{completion_rate}%", total_score])
                
                for col, value in enumerate(values, 1):
                    cell = worksheet.cell(row=row_idx, column=col, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # 交替行颜色
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                        
            except Exception as e:
                print(f"处理问卷 {q['id']} 时出错: {e}")
        
        # 自动调整列宽
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_detail_sheet(self, worksheet, questionnaires):
        """创建详细信息工作表"""
        current_row = 1
        
        for q_idx, q in enumerate(questionnaires):
            try:
                data = json.loads(q['data']) if isinstance(q['data'], str) else q['data']
                
                # 问卷基本信息
                worksheet.cell(row=current_row, column=1, value=f"问卷 #{q['id']}")
                worksheet.cell(row=current_row, column=1).font = Font(bold=True, size=14)
                current_row += 1
                
                basic_info = data.get('basic_info', {})
                info_items = [
                    ('姓名', basic_info.get('name', q['name'])),
                    ('年级', basic_info.get('grade', q['grade'])),
                    ('提交日期', basic_info.get('submission_date', q['submission_date'])),
                    ('问卷类型', q['type'])
                ]
                
                for label, value in info_items:
                    worksheet.cell(row=current_row, column=1, value=label)
                    worksheet.cell(row=current_row, column=2, value=value or '')
                    current_row += 1
                
                current_row += 1  # 空行
                
                # 问题和答案
                questions = data.get('questions', [])
                if questions:
                    worksheet.cell(row=current_row, column=1, value="问题详情")
                    worksheet.cell(row=current_row, column=1).font = Font(bold=True)
                    current_row += 1
                    
                    # 表头
                    headers = ['问题编号', '问题类型', '问题内容', '答案']
                    for col, header in enumerate(headers, 1):
                        cell = worksheet.cell(row=current_row, column=col, value=header)
                        cell.font = Font(bold=True)
                    current_row += 1
                    
                    # 问题数据
                    for question in questions:
                        q_id = question.get('id', '')
                        q_type = question.get('type', '')
                        q_text = question.get('question', '')
                        
                        # 格式化答案
                        answer = self._format_answer_for_excel(question)
                        
                        values = [q_id, q_type, q_text, answer]
                        for col, value in enumerate(values, 1):
                            worksheet.cell(row=current_row, column=col, value=value)
                        current_row += 1
                
                # 统计信息
                stats = data.get('statistics', {})
                if stats:
                    current_row += 1
                    worksheet.cell(row=current_row, column=1, value="统计信息")
                    worksheet.cell(row=current_row, column=1).font = Font(bold=True)
                    current_row += 1
                    
                    for key, value in stats.items():
                        worksheet.cell(row=current_row, column=1, value=key)
                        worksheet.cell(row=current_row, column=2, value=str(value))
                        current_row += 1
                
                # 问卷之间的分隔
                if q_idx < len(questionnaires) - 1:
                    current_row += 2
                    
            except Exception as e:
                print(f"处理问卷详情 {q['id']} 时出错: {e}")
                worksheet.cell(row=current_row, column=1, value=f"问卷 #{q['id']} 处理失败: {str(e)}")
                current_row += 2
        
        # 自动调整列宽
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _format_answer_for_excel(self, question):
        """格式化问题答案用于Excel显示"""
        q_type = question.get('type', '')
        
        if q_type == 'multiple_choice':
            selected = question.get('selected', [])
            options = question.get('options', [])
            
            if not selected:
                return '未选择'
            
            selected_texts = []
            for sel in selected:
                for option in options:
                    if option.get('value') == sel:
                        selected_texts.append(option.get('text', f'选项{sel}'))
                        break
            
            return '; '.join(selected_texts) if selected_texts else '未知选项'
            
        elif q_type == 'text_input':
            return question.get('answer', '未填写')
        
        else:
            # 其他类型，尝试获取通用答案字段
            return str(question.get('answer', question.get('selected', '未知')))
    
    def export_to_pdf(self, questionnaires, include_details=True):
        """导出为PDF格式"""
        buffer = BytesIO()
        
        # 创建PDF文档
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )
        
        # 获取样式
        styles = getSampleStyleSheet()
        
        # 创建自定义样式
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1,  # 居中
            fontName=self.pdf_font
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=12,
            spaceAfter=12,
            fontName=self.pdf_font
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            fontName=self.pdf_font
        )
        
        # 构建PDF内容
        story = []
        
        # 标题
        title = Paragraph("问卷数据报告", title_style)
        story.append(title)
        
        # 生成时间
        generate_time = Paragraph(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style)
        story.append(generate_time)
        story.append(Spacer(1, 12))
        
        # 概览统计
        summary_title = Paragraph("数据概览", heading_style)
        story.append(summary_title)
        
        # 统计表格
        summary_data = [['统计项目', '数值']]
        summary_data.append(['问卷总数', str(len(questionnaires))])
        
        if questionnaires:
            # 按类型统计
            type_counts = {}
            for q in questionnaires:
                q_type = q['type']
                type_counts[q_type] = type_counts.get(q_type, 0) + 1
            
            for q_type, count in type_counts.items():
                summary_data.append([f'{q_type} 类型', str(count)])
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), self.pdf_font),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # 详细数据
        if include_details and questionnaires:
            detail_title = Paragraph("详细数据", heading_style)
            story.append(detail_title)
            
            # 创建详细数据表格
            detail_data = [['ID', '类型', '姓名', '年级', '提交日期']]
            
            for q in questionnaires[:50]:  # 限制PDF中的记录数量，避免文件过大
                try:
                    row = [
                        str(q['id']),
                        q['type'] or '',
                        q['name'] or '',
                        q['grade'] or '',
                        q['submission_date'] or ''
                    ]
                    detail_data.append(row)
                except Exception as e:
                    print(f"处理问卷 {q['id']} 时出错: {e}")
            
            detail_table = Table(detail_data)
            detail_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), self.pdf_font),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8)
            ]))
            
            story.append(detail_table)
            
            # 如果记录数超过50，添加说明
            if len(questionnaires) > 50:
                note = Paragraph(f"注: 由于记录数量较多({len(questionnaires)}条)，PDF中仅显示前50条记录。", normal_style)
                story.append(Spacer(1, 12))
                story.append(note)
        
        # 生成PDF
        doc.build(story)
        
        buffer.seek(0)
        return buffer.getvalue()

# 全局导出器实例
exporter = QuestionnaireExporter()

def export_questionnaires(questionnaires, format_type='csv', include_details=True):
    """
    导出问卷数据
    
    Args:
        questionnaires: 问卷数据列表
        format_type: 导出格式 ('csv', 'excel', 'pdf')
        include_details: 是否包含详细信息
    
    Returns:
        导出的文件内容
    """
    if format_type.lower() == 'csv':
        return exporter.export_to_csv(questionnaires, include_details)
    elif format_type.lower() in ['excel', 'xlsx']:
        return exporter.export_to_excel(questionnaires, include_details)
    elif format_type.lower() == 'pdf':
        return exporter.export_to_pdf(questionnaires, include_details)
    else:
        raise ValueError(f"不支持的导出格式: {format_type}")

def get_export_filename(format_type, prefix='questionnaires'):
    """
    生成导出文件名
    
    Args:
        format_type: 文件格式
        prefix: 文件名前缀
    
    Returns:
        文件名
    """
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    if format_type.lower() == 'csv':
        return f"{prefix}_{timestamp}.csv"
    elif format_type.lower() in ['excel', 'xlsx']:
        return f"{prefix}_{timestamp}.xlsx"
    elif format_type.lower() == 'pdf':
        return f"{prefix}_{timestamp}.pdf"
    else:
        return f"{prefix}_{timestamp}.txt"

def get_content_type(format_type):
    """
    获取文件的Content-Type
    
    Args:
        format_type: 文件格式
    
    Returns:
        Content-Type字符串
    """
    if format_type.lower() == 'csv':
        return 'text/csv; charset=utf-8'
    elif format_type.lower() in ['excel', 'xlsx']:
        return 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    elif format_type.lower() == 'pdf':
        return 'application/pdf'
    else:
        return 'application/octet-stream'